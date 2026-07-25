#!/usr/bin/env python3
"""D1: CQ coverage / completeness matrix.
Introspects the core ontology + SHACL shapes + derived metrics, maps every term to the
competency question(s) that justify it (governance invariant: no term without a CQ), and
records whether each CQ is answerable end-to-end on each built market.
Outputs: evaluation/cq_coverage_matrix.xlsx and evaluation/cq_coverage_matrix.md
"""
import os, glob
from rdflib import Graph, RDF, OWL
from rdflib.namespace import Namespace
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CORE = Namespace("https://w3id.org/ontokg-eq#")

# ---- CQ justification map (term local-name -> CQs). Governance source of truth. ----
CQ = {
 # classes
 "Company":["CQ1","CQ3","CQ4"], "ObservedEntity":["CQ1","CQ2","CQ3","CQ4"],
 "Observation":["CQ1","CQ2","CQ3","CQ4"], "FundamentalObservation":["CQ1"],
 "MarketObservation":["CQ1","CQ3","CQ4"], "ExchangeRateObservation":["CQ2"],
 "ReportingPeriod":["CQ1"], "TemporalInterval":["CQ1","CQ2","CQ3","CQ4"],
 "TimeWindow":["CQ2","CQ3","CQ4"], "AnalysisWindow":["CQ1","CQ2","CQ3"], "EventWindow":["CQ4"],
 "IndustrySectorClassifier":["CQ3"], "IndustrySectorClassificationScheme":["CQ3"],
 "MarketIndex":["CQ1","CQ3"], "Currency":["CQ2"],
 "Announcement":["CQ4"], "Disclosure":["CQ4"], "EvidenceItem":["CQ4","CQ5"],
 "EvidenceSource":["CQ5"], "Publisher":["CQ4","CQ5"], "ProvenanceRecord":["CQ5"],
 "ValidationStatus":["CQ5"], "EvidenceBundle":["CQ5"], "QueryExecution":["CQ5"],
 # object properties
 "isObservationOf":["CQ1","CQ2","CQ3","CQ4"], "observedInPeriod":["CQ1"],
 "observedOverWindow":["CQ1","CQ2","CQ3"], "observedInEventWindow":["CQ4"],
 "hasBaseCurrency":["CQ2"], "hasDealtCurrency":["CQ2"],
 "isClassifiedByIndustrySector":["CQ3"], "isDefinedInIndustrySectorScheme":["CQ3"],
 "usesMarketIndexBenchmark":["CQ1","CQ3"], "aboutCompany":["CQ4"],
 "hasAssociatedDisclosure":["CQ4"], "publishedBy":["CQ4","CQ5"], "hasEvidenceSource":["CQ5"],
 "supportedByEvidenceItem":["CQ5"], "hasProvenanceRecord":["CQ5"], "derivedFrom":["CQ5"],
 "hasValidationStatus":["CQ5"], "containsEvidenceItem":["CQ5"], "includesObservation":["CQ5"],
 "explainsResultEntity":["CQ5"], "isEvidenceBundleFor":["CQ5"], "hasEvidenceBundle":["CQ5"],
 # datatype properties
 "hasCompanyIdentifier":["CQ1","CQ3","CQ4"], "hasCompanyName":["CQ1","CQ3","CQ4"],
 "hasMetricName":["CQ1","CQ2","CQ3","CQ4"], "hasMetricValue":["CQ1","CQ2","CQ3","CQ4"],
 "hasObservationDate":["CQ1","CQ2","CQ3","CQ4"], "hasStartDate":["CQ1","CQ2","CQ3","CQ4"],
 "hasEndDate":["CQ1","CQ2","CQ3","CQ4"], "hasReportDate":["CQ1"], "hasWindowType":["CQ2","CQ3","CQ4"],
 "hasCurrencyCode":["CQ2"], "hasAnnouncementDate":["CQ4"], "hasAnnouncementType":["CQ4"],
 "hasEvidenceIdentifier":["CQ4","CQ5"], "hasEvidenceURL":["CQ4","CQ5"],
 "hasSourceIdentifier":["CQ5"], "hasSourceType":["CQ5"], "hasSourceURL":["CQ5"],
 "hasExtractionDate":["CQ5"], "hasRetrievalDate":["CQ5"], "hasValidationFlag":["CQ5"],
 "hasParameterValues":["CQ5"], "hasQueryFamilyIdentifier":["CQ5"], "hasQueryInstanceIdentifier":["CQ5"],
}
# derived metrics -> CQ
DERIVED = {
 "post-report window return %":["CQ1","CQ3"], "YoY profit growth %":["CQ1"],
 "sector window return %":["CQ3"], "benchmark window return %":["CQ1","CQ3"],
 "FX window change %":["CQ2"], "FX vs company return correlation":["CQ2"],
 "FX vs benchmark return correlation":["CQ2"], "cumulative abnormal return %":["CQ4"],
 "abnormal volume ratio":["CQ4"],
}
# SHACL shapes -> CQ (by target class)
SHAPES = {
 "CompanyShape":["CQ1","CQ3","CQ4"], "ObservationShape":["CQ1","CQ2","CQ3","CQ4"],
 "FundamentalObservationShape":["CQ1"], "MarketObservationShape":["CQ1","CQ3","CQ4"],
 "ExchangeRateObservationShape":["CQ2"], "ReportingPeriodShape":["CQ1"],
 "AnalysisWindowShape":["CQ1","CQ2","CQ3"], "EventWindowShape":["CQ4"],
 "EvidenceItemSourceShape":["CQ5"], "AnnouncementShape":["CQ4"], "EvidenceSourceShape":["CQ5"],
 "ProvenanceRecordShape":["CQ5"], "ValidationStatusShape":["CQ5"], "EvidenceBundleShape":["CQ5"],
 "QueryExecutionShape":["CQ5"],
}

# ---- introspect ontology for completeness check ----
g = Graph().parse(os.path.join(HERE,"ontology","core.ttl"), format="turtle")
def locals_of(t): return sorted(str(s).split("#")[-1] for s in g.subjects(RDF.type, t))
classes = locals_of(OWL.Class)
objprops = locals_of(OWL.ObjectProperty)
dataprops = locals_of(OWL.DatatypeProperty)
all_terms = classes + objprops + dataprops

uncovered = [t for t in all_terms if t not in CQ or not CQ[t]]

# ---- CQ answerability on built markets ----
markets = ["psx","msx"]
cq_rows = {}
for m in markets:
    gm = Graph().parse(os.path.join(HERE,"data",f"demo_{m}.ttl"), format="turtle")
    for q in sorted(glob.glob(os.path.join(HERE,"queries","CQ*_CASE_01.rq"))):
        fam = os.path.basename(q).split("_")[0]
        cq_rows.setdefault(fam,{})[m] = len(list(gm.query(open(q).read())))

# ---- write xlsx ----
wb = openpyxl.Workbook()
hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(color="FFFFFF", bold=True)
def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = Alignment(wrap_text=True, vertical="center")

ws = wb.active; ws.title = "Term-CQ coverage"
ws.append(["Term","Kind","CQ1","CQ2","CQ3","CQ4","CQ5","Justifying CQs"])
def emit(term, kind, cqs):
    flags = ["X" if f"CQ{i}" in cqs else "" for i in range(1,6)]
    ws.append([term, kind]+flags+[", ".join(cqs)])
for t in classes: emit(t,"class",CQ.get(t,[]))
for t in objprops: emit(t,"object property",CQ.get(t,[]))
for t in dataprops: emit(t,"datatype property",CQ.get(t,[]))
for mname,cqs in DERIVED.items(): emit(mname,"derived metric",cqs)
for s,cqs in SHAPES.items(): emit(s,"SHACL shape",cqs)
style_header(ws,8)
ws.column_dimensions['A'].width=40; ws.column_dimensions['B'].width=16; ws.column_dimensions['H'].width=22

ws2 = wb.create_sheet("CQ answerability")
ws2.append(["CQ family","Question (short)","Query template","PSX rows","MSX rows","Answerable"])
short={"CQ1":"fundamentals vs market response","CQ2":"FX context vs market measures",
 "CQ3":"outperformance vs sector & benchmark","CQ4":"announcements vs abnormal activity",
 "CQ5":"explainability & provenance"}
for fam in ["CQ1","CQ2","CQ3","CQ4","CQ5"]:
    p=cq_rows.get(fam,{}).get("psx",0); mx=cq_rows.get(fam,{}).get("msx",0)
    ws2.append([fam, short[fam], f"queries/{fam}_CASE_01.rq", p, mx, "yes (>=1 market)" if (p or mx) else "no"])
style_header(ws2,6)
for col,w in zip("ABCDEF",[10,40,28,10,10,16]): ws2.column_dimensions[col].width=w

out = os.path.join(HERE,"evaluation","cq_coverage_matrix.xlsx")
wb.save(out)

# ---- markdown summary ----
md = os.path.join(HERE,"evaluation","cq_coverage_matrix.md")
with open(md,"w") as f:
    f.write("# CQ coverage / completeness matrix\n\n")
    f.write(f"- Ontology terms: {len(classes)} classes, {len(objprops)} object properties, {len(dataprops)} datatype properties.\n")
    f.write(f"- Derived metrics: {len(DERIVED)}; SHACL shapes: {len(SHAPES)}.\n")
    f.write(f"- **Governance check (every term justified by >=1 CQ): {'PASS' if not uncovered else 'FAIL: '+', '.join(uncovered)}**\n\n")
    f.write("## CQ answerability (rows returned by the computed CQ queries)\n\n")
    f.write("| CQ | Question | PSX | MSX |\n|---|---|---:|---:|\n")
    for fam in ["CQ1","CQ2","CQ3","CQ4","CQ5"]:
        f.write(f"| {fam} | {short[fam]} | {cq_rows.get(fam,{}).get('psx',0)} | {cq_rows.get(fam,{}).get('msx',0)} |\n")

print("classes/objprops/dataprops:", len(classes), len(objprops), len(dataprops))
print("uncovered terms:", uncovered if uncovered else "NONE (all terms CQ-justified)")
print("CQ answerability:", {k:cq_rows[k] for k in sorted(cq_rows)})
print("wrote:", out, "and", md)
